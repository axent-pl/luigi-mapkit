"""XLSX to CSV conversion for Luigi targets."""

import csv
from collections.abc import Callable

import luigi
import openpyxl

RowFilter = Callable[[dict[str, object | None]], bool]


def convert(
    input: luigi.LocalTarget,
    output: luigi.LocalTarget,
    *,
    sheet_name: str | None = None,
    row_filter: RowFilter | None = None,
) -> None:
    """Convert an XLSX worksheet to a CSV file.

    The first worksheet is used when ``sheet_name`` is not provided.
    ``row_filter`` is an optional predicate that receives a row dictionary and
    returns ``True`` if the row should be written to the output CSV.
    """
    workbook = openpyxl.load_workbook(
        input.path,
        read_only=True,
        data_only=True,
    )

    if sheet_name is None:
        if not workbook.sheetnames:
            raise ValueError("workbook contains no worksheets")
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"worksheet {sheet_name!r} not found in workbook"
            )
        worksheet = workbook[sheet_name]

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("worksheet contains no rows") from exc

    if header_row is None:
        raise ValueError("worksheet header row is empty")

    fieldnames = []
    for value in header_row:
        if value is None:
            raise ValueError("worksheet header row must not contain empty column names")
        header_name = str(value).strip()
        if not header_name:
            raise ValueError("worksheet header row must not contain empty column names")
        fieldnames.append(header_name)

    if len(set(fieldnames)) != len(fieldnames):
        raise ValueError("worksheet header row contains duplicate column names")

    with output.open("w") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()

        for row in rows:
            if row is None:
                continue
            row_data = {fieldname: value for fieldname, value in zip(fieldnames, row)}

            if row_filter is not None and not row_filter(row_data):
                continue

            writer.writerow(row_data)
